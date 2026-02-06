"""
Documents API endpoints (US-004, US-005).

Handles document generation, listing, review, and editing.
"""

from datetime import datetime
from typing import Optional

from fastapi import APIRouter, Depends, HTTPException, Query
from fastapi.responses import StreamingResponse
from pydantic import BaseModel
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession

from src.database import get_session
from src.models import DocumentStatus, EnrichedContext, MonthlyDocument, Transaction
from src.services.document_service import DocumentService

router = APIRouter(prefix="/documents", tags=["documents"])


# === Pydantic Models ===


class GenerateDocumentRequest(BaseModel):
    """Request for generating monthly document."""

    user_id: str = "default"
    month: str  # "2026-02"


class GenerateDocumentResponse(BaseModel):
    """Response after document generation."""

    status: str
    document_id: str
    total_transactions: int
    generated_at: str


class DocumentSummary(BaseModel):
    """Document summary for listing."""

    id: str
    month: str
    total_transactions: int
    total_income: int
    total_expense: int
    recurring_count: int
    non_recurring_count: int
    pending_count: int
    status: str
    generated_at: datetime

    class Config:
        from_attributes = True


class DocumentResponse(BaseModel):
    """Full document response."""

    id: str
    user_id: str
    month: str
    total_transactions: int
    total_income: int
    total_expense: int
    recurring_count: int
    non_recurring_count: int
    pending_count: int
    document_markdown: str
    document_version: int
    status: str
    generated_at: datetime
    reviewed_at: Optional[datetime]
    sent_to_accountant_at: Optional[datetime]

    class Config:
        from_attributes = True


class UpdateDocumentRequest(BaseModel):
    """Request for updating document (inline edit)."""

    transaction_id: str
    updates: dict  # {"description": "...", "account_classification": "..."}


class UpdateDocumentResponse(BaseModel):
    """Response after document update."""

    status: str
    document_version: int


class ExcelPreviewRow(BaseModel):
    """Single row for Excel preview."""

    date: str
    amount: int
    counterparty: str
    category: str
    description: str
    account_classification: str
    tax_notes: str


class ExcelPreviewResponse(BaseModel):
    """Excel preview data."""

    rows: list[ExcelPreviewRow]
    summary: dict


# === Helper Functions ===


def document_to_summary(doc: MonthlyDocument) -> DocumentSummary:
    """Convert MonthlyDocument to summary."""
    return DocumentSummary(
        id=doc.id,
        month=doc.month,
        total_transactions=doc.total_transactions,
        total_income=doc.total_income,
        total_expense=doc.total_expense,
        recurring_count=doc.recurring_count,
        non_recurring_count=doc.non_recurring_count,
        pending_count=doc.pending_count,
        status=doc.status.value,
        generated_at=doc.generated_at,
    )


def document_to_response(doc: MonthlyDocument) -> DocumentResponse:
    """Convert MonthlyDocument to response."""
    return DocumentResponse(
        id=doc.id,
        user_id=doc.user_id,
        month=doc.month,
        total_transactions=doc.total_transactions,
        total_income=doc.total_income,
        total_expense=doc.total_expense,
        recurring_count=doc.recurring_count,
        non_recurring_count=doc.non_recurring_count,
        pending_count=doc.pending_count,
        document_markdown=doc.document_markdown or "",
        document_version=doc.document_version,
        status=doc.status.value,
        generated_at=doc.generated_at,
        reviewed_at=doc.reviewed_at,
        sent_to_accountant_at=doc.sent_to_accountant_at,
    )


# === Endpoints ===


@router.post("/generate", response_model=GenerateDocumentResponse)
async def generate_document(
    request: GenerateDocumentRequest,
    session: AsyncSession = Depends(get_session),
):
    """
    Generate monthly document.

    Collects all transactions and enriched contexts for the month,
    then generates a comprehensive markdown document.

    AC-004-01, AC-004-02, AC-004-03, AC-004-04
    """
    # Parse month
    try:
        year, month = request.month.split("-")
        year = int(year)
        month = int(month)
    except ValueError:
        raise HTTPException(status_code=400, detail="Invalid month format. Use YYYY-MM")

    if not (1 <= month <= 12):
        raise HTTPException(status_code=400, detail="Month must be between 1 and 12")

    # Generate document
    service = DocumentService(session)
    document = await service.generate_monthly_document(
        user_id=request.user_id,
        year=year,
        month=month,
    )

    return GenerateDocumentResponse(
        status="success",
        document_id=document.id,
        total_transactions=document.total_transactions,
        generated_at=document.generated_at.isoformat(),
    )


@router.get("/", response_model=list[DocumentSummary])
async def list_documents(
    user_id: str = Query("default", description="User ID"),
    year: Optional[int] = Query(None, description="Filter by year"),
    session: AsyncSession = Depends(get_session),
):
    """
    List all documents for a user.

    Optionally filter by year.
    """
    query = select(MonthlyDocument).where(MonthlyDocument.user_id == user_id)

    if year:
        query = query.where(MonthlyDocument.month.startswith(str(year)))

    query = query.order_by(MonthlyDocument.month.desc())

    result = await session.execute(query)
    documents = result.scalars().all()

    return [document_to_summary(doc) for doc in documents]


@router.get("/{document_id}", response_model=DocumentResponse)
async def get_document(
    document_id: str,
    session: AsyncSession = Depends(get_session),
):
    """
    Get a single document by ID.
    """
    document = await session.get(MonthlyDocument, document_id)

    if not document:
        raise HTTPException(status_code=404, detail="Document not found")

    return document_to_response(document)


@router.put("/{document_id}", response_model=UpdateDocumentResponse)
async def update_document(
    document_id: str,
    request: UpdateDocumentRequest,
    session: AsyncSession = Depends(get_session),
):
    """
    Update document (inline edit).

    Updates the EnrichedContext for the specified transaction
    and regenerates the document.

    AC-005-03
    """
    document = await session.get(MonthlyDocument, document_id)
    if not document:
        raise HTTPException(status_code=404, detail="Document not found")

    # Get transaction and context
    transaction = await session.get(Transaction, request.transaction_id)
    if not transaction:
        raise HTTPException(status_code=404, detail="Transaction not found")

    # Get or create EnrichedContext
    context_query = select(EnrichedContext).where(
        EnrichedContext.transaction_id == request.transaction_id
    )
    context_result = await session.execute(context_query)
    context = context_result.scalar_one_or_none()

    if context:
        # Update existing context
        for key, value in request.updates.items():
            if hasattr(context, key):
                setattr(context, key, value)
    else:
        # Create new context with updates
        import uuid
        context = EnrichedContext(
            id=f"EC-{transaction.date.strftime('%Y-%m-%d')}-{uuid.uuid4().hex[:6]}",
            transaction_id=transaction.id,
            **{k: v for k, v in request.updates.items() if hasattr(EnrichedContext, k)}
        )
        session.add(context)

    # Regenerate document
    year, month = document.month.split("-")
    service = DocumentService(session)
    updated_doc = await service.generate_monthly_document(
        user_id=document.user_id,
        year=int(year),
        month=int(month),
    )

    return UpdateDocumentResponse(
        status="success",
        document_version=updated_doc.document_version,
    )


@router.post("/{document_id}/review")
async def mark_reviewed(
    document_id: str,
    session: AsyncSession = Depends(get_session),
):
    """
    Mark document as reviewed.

    AC-005-06
    """
    document = await session.get(MonthlyDocument, document_id)
    if not document:
        raise HTTPException(status_code=404, detail="Document not found")

    document.status = DocumentStatus.REVIEWED
    document.reviewed_at = datetime.utcnow()

    await session.commit()

    return {"status": "success", "reviewed_at": document.reviewed_at.isoformat()}


@router.get("/{document_id}/excel-preview", response_model=ExcelPreviewResponse)
async def excel_preview(
    document_id: str,
    session: AsyncSession = Depends(get_session),
):
    """
    Get Excel preview data.

    Returns structured data for Excel preview display.

    AC-005-05
    """
    document = await session.get(MonthlyDocument, document_id)
    if not document:
        raise HTTPException(status_code=404, detail="Document not found")

    # Get transactions for the month
    year, month = document.month.split("-")
    start_date = datetime(int(year), int(month), 1)
    if int(month) == 12:
        end_date = datetime(int(year) + 1, 1, 1)
    else:
        end_date = datetime(int(year), int(month) + 1, 1)

    tx_query = (
        select(Transaction)
        .where(Transaction.date >= start_date)
        .where(Transaction.date < end_date)
        .where(Transaction.is_internal_transfer == False)  # noqa: E712
        .order_by(Transaction.date)
    )
    tx_result = await session.execute(tx_query)
    transactions = tx_result.scalars().all()

    # Get contexts
    tx_ids = [tx.id for tx in transactions]
    ctx_query = select(EnrichedContext).where(EnrichedContext.transaction_id.in_(tx_ids))
    ctx_result = await session.execute(ctx_query)
    contexts = {ctx.transaction_id: ctx for ctx in ctx_result.scalars().all()}

    # Build preview rows
    rows = []
    for tx in transactions:
        ctx = contexts.get(tx.id)
        sign = -1 if tx.type.value == "지출" else 1

        rows.append(ExcelPreviewRow(
            date=tx.date.strftime("%Y-%m-%d"),
            amount=tx.amount * sign,
            counterparty=tx.counterparty or "",
            category="정기" if tx.is_recurring else "비정기",
            description=ctx.ai_generated_summary if ctx else tx.bank_memo or "",
            account_classification=ctx.account_classification if ctx else "",
            tax_notes=ctx.tax_notes if ctx else "",
        ))

    return ExcelPreviewResponse(
        rows=rows,
        summary={
            "total_transactions": document.total_transactions,
            "total_income": document.total_income,
            "total_expense": document.total_expense,
            "net_flow": document.total_income - document.total_expense,
        },
    )


@router.get("/{document_id}/download")
async def download_excel(
    document_id: str,
    session: AsyncSession = Depends(get_session),
):
    """
    Download document as Excel file.

    AC-006-02
    """
    document = await session.get(MonthlyDocument, document_id)
    if not document:
        raise HTTPException(status_code=404, detail="Document not found")

    # Get preview data
    preview = await excel_preview(document_id, session)

    # Generate Excel using openpyxl
    from io import BytesIO
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill

    wb = Workbook()

    # Sheet 1: Transaction Details
    ws1 = wb.active
    ws1.title = "거래 내역"

    # Header
    headers = ["날짜", "금액", "거래처", "구분", "설명", "계정 분류", "세무 처리"]
    for col, header in enumerate(headers, 1):
        cell = ws1.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="CCE5FF", end_color="CCE5FF", fill_type="solid")

    # Data
    for row_idx, row in enumerate(preview.rows, 2):
        ws1.cell(row=row_idx, column=1, value=row.date)
        ws1.cell(row=row_idx, column=2, value=row.amount)
        ws1.cell(row=row_idx, column=3, value=row.counterparty)
        ws1.cell(row=row_idx, column=4, value=row.category)
        ws1.cell(row=row_idx, column=5, value=row.description)
        ws1.cell(row=row_idx, column=6, value=row.account_classification)
        ws1.cell(row=row_idx, column=7, value=row.tax_notes)

    # Adjust column widths
    ws1.column_dimensions["A"].width = 12
    ws1.column_dimensions["B"].width = 15
    ws1.column_dimensions["C"].width = 20
    ws1.column_dimensions["D"].width = 10
    ws1.column_dimensions["E"].width = 40
    ws1.column_dimensions["F"].width = 20
    ws1.column_dimensions["G"].width = 30

    # Sheet 2: Summary
    ws2 = wb.create_sheet("요약")
    ws2.cell(row=1, column=1, value="항목").font = Font(bold=True)
    ws2.cell(row=1, column=2, value="금액").font = Font(bold=True)

    ws2.cell(row=2, column=1, value="총 입금")
    ws2.cell(row=2, column=2, value=preview.summary["total_income"])

    ws2.cell(row=3, column=1, value="총 지출")
    ws2.cell(row=3, column=2, value=preview.summary["total_expense"])

    ws2.cell(row=4, column=1, value="순 현금흐름")
    ws2.cell(row=4, column=2, value=preview.summary["net_flow"])

    ws2.cell(row=5, column=1, value="총 거래 건수")
    ws2.cell(row=5, column=2, value=preview.summary["total_transactions"])

    # Save to buffer
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    # Generate filename
    filename = f"{document.month}_입출금내역.xlsx"

    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )
