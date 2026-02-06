"""
Delivery API endpoints (US-006).

Handles sending documents to accountant via email.
"""

from datetime import datetime
from io import BytesIO
from typing import Optional

from fastapi import APIRouter, Depends, HTTPException
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from pydantic import BaseModel, EmailStr
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession

from src.database import get_session
from src.models import DocumentStatus, EnrichedContext, MonthlyDocument, Transaction
from src.services.email_service import EmailService

router = APIRouter(prefix="/delivery", tags=["delivery"])


# === Pydantic Models ===


class SendToAccountantRequest(BaseModel):
    """Request for sending document to accountant."""

    document_id: str
    accountant_email: str
    format: str = "xlsx"  # xlsx, csv, pdf
    user_name: str = "사용자"


class SendToAccountantResponse(BaseModel):
    """Response after sending to accountant."""

    status: str
    sent_at: Optional[str] = None
    error: Optional[str] = None


class DeliveryStatusResponse(BaseModel):
    """Delivery status response."""

    document_id: str
    status: str
    sent_to_accountant_at: Optional[str] = None
    accountant_email: Optional[str] = None


# === Endpoints ===


@router.post("/send", response_model=SendToAccountantResponse)
async def send_to_accountant(
    request: SendToAccountantRequest,
    session: AsyncSession = Depends(get_session),
):
    """
    Send document to accountant via email.

    Generates Excel file and sends via SMTP.

    AC-006-01, AC-006-02, AC-006-03
    """
    # Get document
    document = await session.get(MonthlyDocument, request.document_id)
    if not document:
        raise HTTPException(status_code=404, detail="Document not found")

    # Check if document is reviewed
    if document.status == DocumentStatus.GENERATED:
        raise HTTPException(
            status_code=400,
            detail="Document must be reviewed before sending. Please review first.",
        )

    # Get transactions for Excel
    year, month = document.month.split("-")
    start_date = datetime(int(year), int(month), 1)
    if int(month) == 12:
        end_date = datetime(int(year) + 1, 1, 1)
    else:
        end_date = datetime(int(year), int(month) + 1, 1)

    tx_query = (
        select(Transaction)
        .where(Transaction.date >= start_date)
        .where(Transaction.date < end_date)
        .where(Transaction.is_internal_transfer == False)  # noqa: E712
        .order_by(Transaction.date)
    )
    tx_result = await session.execute(tx_query)
    transactions = tx_result.scalars().all()

    # Get contexts
    tx_ids = [tx.id for tx in transactions]
    ctx_query = select(EnrichedContext).where(EnrichedContext.transaction_id.in_(tx_ids))
    ctx_result = await session.execute(ctx_query)
    contexts = {ctx.transaction_id: ctx for ctx in ctx_result.scalars().all()}

    # Generate Excel
    excel_content = _generate_excel(document, transactions, contexts)
    filename = f"{request.user_name}_{document.month}_입출금내역.xlsx"

    # Send email
    email_service = EmailService()
    result = await email_service.send_to_accountant(
        to_email=request.accountant_email,
        user_name=request.user_name,
        month=document.month,
        total_transactions=document.total_transactions,
        total_income=document.total_income,
        total_expense=document.total_expense,
        attachment_content=excel_content,
        attachment_filename=filename,
    )

    # Update document status
    if result["status"] in ["sent", "mock_sent"]:
        document.status = DocumentStatus.SENT
        document.sent_to_accountant_at = datetime.utcnow()
        document.accountant_email = request.accountant_email
        await session.commit()

        return SendToAccountantResponse(
            status="success",
            sent_at=result.get("sent_at"),
        )
    else:
        return SendToAccountantResponse(
            status="error",
            error=result.get("error", "Unknown error"),
        )


@router.get("/status/{document_id}", response_model=DeliveryStatusResponse)
async def get_delivery_status(
    document_id: str,
    session: AsyncSession = Depends(get_session),
):
    """
    Get delivery status for a document.

    AC-006-04
    """
    document = await session.get(MonthlyDocument, document_id)
    if not document:
        raise HTTPException(status_code=404, detail="Document not found")

    return DeliveryStatusResponse(
        document_id=document.id,
        status=document.status.value,
        sent_to_accountant_at=(
            document.sent_to_accountant_at.isoformat() if document.sent_to_accountant_at else None
        ),
        accountant_email=document.accountant_email,
    )


# === Helper Functions ===


def _generate_excel(
    document: MonthlyDocument,
    transactions: list[Transaction],
    contexts: dict[str, EnrichedContext],
) -> bytes:
    """Generate Excel file for accountant delivery."""
    wb = Workbook()

    # Sheet 1: Transaction Details
    ws1 = wb.active
    ws1.title = "거래 내역"

    # Header
    headers = ["날짜", "금액", "거래처", "구분", "설명", "계정 분류", "세무 처리"]
    for col, header in enumerate(headers, 1):
        cell = ws1.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="CCE5FF", end_color="CCE5FF", fill_type="solid")

    # Data
    for row_idx, tx in enumerate(transactions, 2):
        ctx = contexts.get(tx.id)
        sign = -1 if tx.type.value == "지출" else 1

        ws1.cell(row=row_idx, column=1, value=tx.date.strftime("%Y-%m-%d"))
        ws1.cell(row=row_idx, column=2, value=tx.amount * sign)
        ws1.cell(row=row_idx, column=3, value=tx.counterparty or "")
        ws1.cell(row=row_idx, column=4, value="정기" if tx.is_recurring else "비정기")
        ws1.cell(row=row_idx, column=5, value=ctx.ai_generated_summary if ctx else tx.bank_memo or "")
        ws1.cell(row=row_idx, column=6, value=ctx.account_classification if ctx else "")
        ws1.cell(row=row_idx, column=7, value=ctx.tax_notes if ctx else "")

    # Adjust column widths
    ws1.column_dimensions["A"].width = 12
    ws1.column_dimensions["B"].width = 15
    ws1.column_dimensions["C"].width = 20
    ws1.column_dimensions["D"].width = 10
    ws1.column_dimensions["E"].width = 40
    ws1.column_dimensions["F"].width = 20
    ws1.column_dimensions["G"].width = 30

    # Sheet 2: Summary
    ws2 = wb.create_sheet("요약")
    ws2.cell(row=1, column=1, value="항목").font = Font(bold=True)
    ws2.cell(row=1, column=2, value="금액").font = Font(bold=True)

    ws2.cell(row=2, column=1, value="총 입금")
    ws2.cell(row=2, column=2, value=document.total_income)

    ws2.cell(row=3, column=1, value="총 지출")
    ws2.cell(row=3, column=2, value=document.total_expense)

    ws2.cell(row=4, column=1, value="순 현금흐름")
    ws2.cell(row=4, column=2, value=document.total_income - document.total_expense)

    ws2.cell(row=5, column=1, value="총 거래 건수")
    ws2.cell(row=5, column=2, value=document.total_transactions)

    # Sheet 3: Tax Notes
    ws3 = wb.create_sheet("세무 처리 메모")
    ws3.cell(row=1, column=1, value="거래처").font = Font(bold=True)
    ws3.cell(row=1, column=2, value="금액").font = Font(bold=True)
    ws3.cell(row=1, column=3, value="세무 처리 메모").font = Font(bold=True)

    row_idx = 2
    for tx in transactions:
        ctx = contexts.get(tx.id)
        if ctx and ctx.tax_notes:
            ws3.cell(row=row_idx, column=1, value=tx.counterparty or "")
            ws3.cell(row=row_idx, column=2, value=tx.amount)
            ws3.cell(row=row_idx, column=3, value=ctx.tax_notes)
            row_idx += 1

    # Save to buffer
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    return buffer.read()
